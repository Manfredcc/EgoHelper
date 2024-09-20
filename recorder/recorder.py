import tkinter as tk
import pandas as pd
import pyperclip
import openpyxl
import time
import threading
import logging
from fuzzywuzzy import process
from fuzzywuzzy import fuzz

# === 配置开始 ===
name = '记录助手'
independent = 1 # 0：作为模块，1：单独使用
recordPath = 'record.xlsx' # 保存数据的默认路径
listBoxLength = 5 # 每次抓取记录数据的条目数量
windowSize = [500, listBoxLength * 20 + 200]
logging.basicConfig(level=logging.DEBUG)
# logging.disable()
# === 配置结束 ===

gThreadPool = [] # 线程池
gStopEvent = threading.Event() # 线程退出标志位

# 将屏幕置于当前显示器中央
def windowCenter(window, size):
    screenWidth = window.winfo_screenwidth()
    screenHeight = window.winfo_screenheight()
    x = (screenWidth // 2) - (size[0] // 2)
    y = (screenHeight // 2) - (size[1] // 2)
    # 将屏幕左上角放置在坐标(x,y)处，面积为(width x height)
    window.geometry(f"{size[0]}x{size[1]}+{int(x)}+{int(y)}")

# 保存键值对
def store(filename, key, val):
    try:
        wb = openpyxl.load_workbook(filename)
    except FileNotFoundError:
        wb = openpyxl.Workbook()
    wSheet = wb.active
    wSheet.append([key, val])
    wb.save(filename)

# 将数据记录到recordPath中，并将记录信息通过echo标签回显
def record(eKey, eVal, lEcho):
    # 保存数据
    val = eVal.get().strip()
    key = eKey.get().strip()
    if val == '': # 输入框无数据，则记录剪切板数据
        val = pyperclip.paste()
    store(recordPath, key, val)
    # 回显保存数据
    if key != '':
        lEcho.config(text=key+' : '+ val)
    else:
        lEcho.config(text=val)

# 读取start~start+listBoxLength行的数据，并分别计算匹配度，
# 只保留匹配度最高的前listBoxLength个数据
#
# 模糊匹配可选算法
# fuzz.partial_ratio：只考虑最长公共子序列长度
# fuzz.token_sort_ratio：将字符串分割成词（基于空格），然后对每个词进行排序，比较排序后的列表
# fuzz.token_set_ratio：将字符串分割成词，然后比较两个字符串中相同词的数量
def getRowsByRation(ws, start, tarStr, topStr):
    data = []
    rationStr = []
    for row in ws.iter_rows(min_row=start, max_row=start+listBoxLength-1, values_only=True):
        cellVal = row[:2]
        data.append(cellVal)
    for s in data:
        ration = fuzz.partial_ratio(tarStr, s)
        rationStr.append({'str': s, 'ratio': ration})
    topStr.extend(rationStr)
    topStr.sort(key=lambda x:x['ratio'], reverse=True)
    topStr = topStr[:listBoxLength]

# 搜索记录，将匹配度最高的listBoxLength个数据，展示在搜索结果展列表
def match(filepath, target, listBoxVar):
    try:
        wb = openpyxl.load_workbook(filepath)
    except FileNotFoundError:
        wb = openpyxl.Workbook()
    wSheet = wb.active
    topRationStr = []
    i = 1
    while i <= wSheet.max_row:
        getRowsByRation(wSheet, i, target, topRationStr)
        filerStr = [item['str'] for item in topRationStr if item['str'] != (None, None)]
        listBoxVar.set(filerStr)
        time.sleep(0.1)
        if gStopEvent.is_set():
            logging.info('match exit')
            return
        i += listBoxLength

def showMatch(target, listBoxVar):
    oldTarstr = ''
    while True:
        tarStr = target.get().strip()
        if (tarStr != '' and oldTarstr != tarStr): # 关键词输入框有新数据
            oldTarstr = tarStr
            match(recordPath, tarStr, listBoxVar)
        time.sleep(0.2)
        if gStopEvent.is_set():
            logging.info('showMatch exit')
            return

def onListSelect(event):
    listbox = event.widget
    selectedIndex = listbox.curselection()
    if not selectedIndex:
        return
    selectedItem = listbox.get(selectedIndex[0])
    pyperclip.copy(selectedItem[1]) # 将key-val的val复制到剪切板
    logging.debug(f"你选择了:{selectedItem}")

def onClosing(window):
    gStopEvent.set()
    for thread in gThreadPool:
        try:
            thread.join(timeout=1)
        except Exception as e:
            logging.error('等待线程 %s 退出时发生错误', thread.name, str(e))
    window.destroy()
    logging.info('窗口关闭')

def main(window):
    # === 添加记录界面 -start- ===
    fRecordKeyVal = tk.Frame(window)
    # 标签：回显添加记录的数据
    lEcho = tk.Label(fRecordKeyVal, wraplength=400)
    # 输入框：键值对（可选输入）
    eKey = tk.Entry(fRecordKeyVal, width=40)
    lKey = tk.Label(fRecordKeyVal, text='关键词（可选）')
    eVal = tk.Entry(fRecordKeyVal, width=40)
    lVal = tk.Label(fRecordKeyVal, text='值（空则代表使用剪切板）')
    #记录按钮
    bRecord = tk.Button(fRecordKeyVal, text='点此记录', bg='orange', command=
                            lambda: record(eKey, eVal, lEcho))
    # 布局
    fRecordKeyVal.grid(row=0, column=0, sticky='w', padx=40)
    eKey.grid(row=0, column=0, sticky='w')
    lKey.grid(row=0, column=1, sticky='w')
    eVal.grid(row=1, sticky='w')
    lVal.grid(row=1, column=1, sticky='w')
    bRecord.grid(row=0, column=1, sticky='e')
    lEcho.grid(row=3, column=0, sticky='w')
    # === 添加记录界面 -end- ===

    # === 添加搜索界面 -start- ===
    fFuzzySearch = tk.Frame(window)
    # 搜索关键词输入框
    eTarget = tk.Entry(fFuzzySearch, width=40)
    lTarget = tk.Label(fFuzzySearch, text='匹配词')
    # 搜索结果展列表
    listBoxVar = tk.StringVar()
    lbMatch = tk.Listbox(fFuzzySearch, listvariable=listBoxVar, width=40, height=listBoxLength)
    # 布局
    fFuzzySearch.grid(row=1, column=0, sticky='w', padx=40, pady=20)
    eTarget.grid(row=0, column=0, sticky='w')
    lTarget.grid(row=0, column=1, sticky='e')
    lbMatch.grid(row=1, column=0, sticky='w')
    # === 添加搜索界面 -end- ===

    # 启动匹配线程，模糊匹配关键词输入框的key，并展示结果在搜索列表
    threadMatch = threading.Thread(target=showMatch, args=(eTarget, listBoxVar))
    gThreadPool.append(threadMatch)
    threadMatch.start()
    # 点击候选列表，即复制内容到剪切板
    lbMatch.bind("<<ListboxSelect>>", onListSelect)
    # 关闭窗口时，退出所有线程
    window.protocol("WM_DELETE_WINDOW", lambda: onClosing(window))
    # 主界面循环
    window.mainloop()


if independent:
    window = tk.Tk()
    window.title('记录助手')
    windowCenter(window, windowSize)
    main(window)